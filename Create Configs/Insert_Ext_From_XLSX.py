# -*- coding: utf-8 -*-
# !/usr/bin/env python

from openpyxl import load_workbook
import csv
import MySQLdb
conn = MySQLdb.connect(host= "localhost",
                  user="root",
                  passwd="PASSWORD",
                  db="Prod6",
                  charset="utf8"
                 )
x = conn.cursor()

wb = load_workbook('ImportPhoneBookRUS.xlsx', data_only=True)
sheet = wb['Sheet']

PhonesDict = {}

for row in sheet.iter_rows():
     #print(row[1].value)
     name = row[2].value
     extension = row[1].value
     secret = row[8].value
     callgroup = row[4].value
     pickup = row[5].value
     mac = row[3].value
     ring = row[6].value
     email = row[7].value
     ip3 = row[9].value
     ip4 = row[10].value
     model = row[11].value
     DID = row[12].value
     #DID1 = ('"'+str(extension)+" "+name+'"'+" "+'<'+str(DID)+'>')
     #print('"'+str(extension)+" "+name+'"'+" "+'<'+str(DID)+'>')
     #DID = DID1
     print(DID)
     x.execute("INSERT INTO users(name,ext,secret,callGroup,pickup,mac,ring,email,ip3,ip4,model,DID) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",(name, extension, secret, callgroup, pickup, mac, ring, email, ip3, ip4, model, DID))
     conn.commit()
conn.close()
     #for cell in row:
     #     print(cell.value)
